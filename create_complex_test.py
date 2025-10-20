"""
複雑なテストフローチャートExcelファイルを生成するスクリプト

仕様:
- 開始ノード（角丸四角形）
- 処理ノード（四角形）
- 分岐ノード（ひし形）
- Yes/Noラベル（矢印の横のテキストボックス）
- 終了ノード（角丸四角形）
"""

import openpyxl
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, EMU_to_pixels
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.drawing.geometry import PresetGeometry2D, ShapeStyle
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import RGBColor
import openpyxl.drawing.shapes as shapes

def create_complex_flowchart():
    """複雑なフローチャートを含むExcelファイルを生成"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    print("⚠️  注意: openpyxlでは図形の完全な作成が制限されています。")
    print("   このスクリプトは基本的なセル構造のみを作成します。")
    print("   実際の図形は手動でExcelで追加する必要があります。")
    print()
    print("以下の図形を手動で追加してください:")
    print()
    print("1. セルB2付近: 角丸四角形 - テキスト「開始」")
    print("2. セルB4付近: 四角形 - テキスト「データ入力」")
    print("3. セルB6付近: ひし形（分岐） - テキスト「データ有効？」")
    print("4. セルD8付近: 四角形 - テキスト「エラー表示」")
    print("5. セルB8付近: 四角形 - テキスト「データ処理」")
    print("6. セルB10付近: 角丸四角形 - テキスト「終了」")
    print()
    print("矢印とラベル:")
    print("- 開始 → データ入力")
    print("- データ入力 → データ有効？")
    print("- データ有効？ → データ処理 (矢印の横にテキストボックス「Yes」)")
    print("- データ有効？ → エラー表示 (矢印の横にテキストボックス「No」)")
    print("- データ処理 → 終了")
    print("- エラー表示 → 終了")

    # ガイド用のテキストをセルに配置
    ws['B2'] = '【ここに角丸四角形: 開始】'
    ws['B4'] = '【ここに四角形: データ入力】'
    ws['B6'] = '【ここにひし形: データ有効？】'
    ws['B8'] = '【ここに四角形: データ処理】'
    ws['D8'] = '【ここに四角形: エラー表示】'
    ws['B10'] = '【ここに角丸四角形: 終了】'

    ws['C6'] = '← Yes'
    ws['E6'] = '← No'

    # 列幅を調整
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15

    # 行の高さを調整
    for row in [2, 4, 6, 8, 10]:
        ws.row_dimensions[row].height = 40

    wb.save('test_complex_chart_template.xlsx')
    print()
    print("✓ test_complex_chart_template.xlsx を作成しました。")
    print("  このファイルをExcelで開き、上記の指示に従って図形を手動で追加してください。")
    print("  完成したら test_complex_chart.xlsx として保存してください。")

if __name__ == '__main__':
    create_complex_flowchart()
